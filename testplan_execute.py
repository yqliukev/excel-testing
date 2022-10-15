import command_execution as exe
import openpyxl as xl
import pytest
import pandas as pd

# command run function for each run
# Arguments:
    # active sheet as sh
    # row number as line 
def runCommand(sh, line):
    # set base command in first column, encased in quotes, output and err of last command will be returned
    cmds = sh.cell(row = line, column = 1)

    # set options in second column, enclosed in quotes
    options = sh.cell(row = line, column = 2).value.split('"')[1::2]
    for option in options:
        cmd = cmd + " " + option
    out, err = exe.execute_shell_command(cmd)
    return out, err



def runTest(fileName):

    wrkbk = xl.load_workbook(fileName)
    sh = wrkbk.active

    # run output check column 
    # output column should have python check without assert assuming command has been run
    for i in range(2, sh.max_row + 1):
        out, err = runCommand(sh, i)
        assertcommand = "assert "

        assertcommand = assertcommand + sh.cell(row = i, column = 3).value
        exec(assertcommand)
    
        


runTest("Tests/Basic Test.xlsx")

